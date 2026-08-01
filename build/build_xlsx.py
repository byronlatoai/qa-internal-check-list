# -*- coding: utf-8 -*-
"""
Genera la plantilla en Excel a partir de data/checklist.json.

Una pestaña por documento, autocontenida, con este orden interno:
    1) TRANSVERSALES        (inyectados desde transversales.json)
    2) CRÍTICOS DE ...      (los críticos propios de ese documento)
    3) el resto, agrupado por mini-categoría

Correr antes build/build_data.py si se tocaron las fuentes.
Uso:   python build/build_xlsx.py
"""

import io, os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data", "checklist.json")
XLSX = os.path.join(ROOT, "downloads", "QA_Checklist_LATO.xlsx")

D = json.load(io.open(DATA, encoding="utf-8"))

ARIAL     = "Arial"
INK       = "1F3864"
HDRFILL   = PatternFill("solid", fgColor="1F3864")
BLOCKFILL = PatternFill("solid", fgColor="DCE3EF")
TRANSFILL = PatternFill("solid", fgColor="E7F0E9")
CRITFILL  = PatternFill("solid", fgColor="FDECEF")
BANDFILL  = PatternFill("solid", fgColor="F7F8FA")
LEGFILL   = PatternFill("solid", fgColor="FFF2CC")
thin      = Side(style="thin", color="D0D5DD")
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

HEAD   = ["#", "Categoría", "Crítico", "Qué revisar", "Por qué / de dónde sale",
          "¿Va?", "Notas / decisión"]
WIDTHS = [8, 18, 11, 76, 40, 12, 38]
DV     = '"Sí,No,Ajustar,Pendiente"'

wb = Workbook()
cover = wb.active
cover.title = "Instrucciones"
resumen = []

for idx, doc in enumerate(D["documentos"]):
    nombre = ("%d. %s" % (idx + 1, doc["corto"]))[:31]   # Excel corta en 31
    ws = wb.create_sheet(nombre)

    ws["A1"] = doc["titulo"]
    ws["A1"].font = Font(name=ARIAL, size=13, bold=True, color=INK)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.row_dimensions[1].height = 22

    ws["A2"] = ("%d ítems, %d críticos.  %s  Marca en «¿Va?» y anota en «Notas / decisión»."
                % (doc["total"], doc["criticos"], doc["cuando"]))
    ws["A2"].font = Font(name=ARIAL, size=9, italic=True, color="667085")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    for c, h in enumerate(HEAD, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        cell.fill = HDRFILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[4].height = 30

    r = 5
    filas_dv = []
    for b in doc["bloques"]:
        if not b["items"]:
            continue
        etiqueta = b["titulo"].upper()
        if b["subtitulo"]:
            etiqueta += "  —  " + b["subtitulo"]
        ws.cell(row=r, column=1, value="%s  ·  %d" % (etiqueta, len(b["items"])))
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.fill = TRANSFILL if b["tipo"] == "transversal" else BLOCKFILL
            cell.font = Font(name=ARIAL, size=10, bold=True, color=INK)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 20
        r += 1

        for k, it in enumerate(b["items"]):
            vals = [it["id"], it["categoria"], "CRÍTICO" if it["critico"] else "",
                    it["texto"], it["fuente"], "", ""]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = Font(name=ARIAL, size=10,
                                 bold=(it["critico"] and c in (1, 3)),
                                 color="B42318" if (c == 3 and it["critico"]) else "101828")
                cell.alignment = Alignment(vertical="top",
                                           wrap_text=(c in (2, 4, 5, 7)),
                                           horizontal="center" if c in (1, 3, 6) else "left")
                cell.border = BORDER
                if it["critico"]:
                    cell.fill = CRITFILL
                elif k % 2:
                    cell.fill = BANDFILL
            filas_dv.append(r)
            r += 1

    dv = DataValidation(type="list", formula1=DV, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    for f in filas_dv:
        dv.add("F%d" % f)

    for c, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    resumen.append((nombre, doc["total"], doc["criticos"], r - 1))

# ------------------------------------------------ pestaña fuente transversales
trans = D["documentos"][0]["bloques"][0]["items"]
wt = wb.create_sheet("T. Transversales (fuente)")
wt["A1"] = "TRANSVERSALES — fuente única"
wt["A1"].font = Font(name=ARIAL, size=13, bold=True, color=INK)
wt.merge_cells("A1:E1")
wt["A2"] = ("Estas %d reglas se repiten como primer bloque en cada pestaña de documento. "
            "Se editan en data/transversales.json y se regeneran con build/build_data.py "
            "y build/build_xlsx.py." % len(trans))
wt["A2"].font = Font(name=ARIAL, size=9, italic=True, color="667085")
wt.merge_cells("A2:E2")
for c, h in enumerate(["#", "Categoría", "Crítico", "Qué revisar", "Por qué / de dónde sale"], start=1):
    cell = wt.cell(row=4, column=c, value=h)
    cell.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    cell.fill = HDRFILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
wt.row_dimensions[4].height = 30
for k, it in enumerate(trans):
    r = 5 + k
    for c, v in enumerate([it["id"], it["categoria"], "CRÍTICO" if it["critico"] else "",
                           it["texto"], it["fuente"]], start=1):
        cell = wt.cell(row=r, column=c, value=v)
        cell.font = Font(name=ARIAL, size=10, bold=(it["critico"] and c in (1, 3)),
                         color="B42318" if (c == 3 and it["critico"]) else "101828")
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4, 5)),
                                   horizontal="center" if c in (1, 3) else "left")
        cell.border = BORDER
        if it["critico"]:
            cell.fill = CRITFILL
        elif k % 2:
            cell.fill = BANDFILL
for c, w in enumerate([8, 18, 11, 76, 40], start=1):
    wt.column_dimensions[get_column_letter(c)].width = w
wt.freeze_panes = "A5"
wt.sheet_view.showGridLines = False

# ------------------------------------------------------------------ portada
cover.sheet_view.showGridLines = False
cover["A1"] = "QA CHECKLIST — LATO AI"
cover["A1"].font = Font(name=ARIAL, size=18, bold=True, color=INK)
cover["A2"] = "Plantilla de revisión · contenido %s" % D["contenido"]
cover["A2"].font = Font(name=ARIAL, size=11, color="667085")

cover["A4"] = "Fuente"
cover["A4"].font = Font(name=ARIAL, size=10, bold=True)
cover["A5"] = ("Feedback de Dana Carvajal en los 20 chats de proyecto de Telegram (abril–julio 2026), "
               "sus 16 videos de revisión, sus 41 notas de voz transcritas, las capturas anotadas que envió, "
               "la Knowledge Base del equipo, y la comparación entre la proposal generada por el sistema "
               "para Baily y la versión que ella corrigió a mano.")
cover["A5"].font = Font(name=ARIAL, size=10)
cover["A5"].alignment = Alignment(wrap_text=True, vertical="top")
cover.merge_cells("A5:E7")

cover["A9"] = "Cómo se usa"
cover["A9"].font = Font(name=ARIAL, size=10, bold=True)
for i, t in enumerate([
    "1.  Una pestaña por documento. Cada pestaña es autocontenida: trae TODO lo que hay que revisar",
    "     en ese entregable, incluidos los transversales.",
    "2.  Dentro de cada pestaña el orden es: TRANSVERSALES, luego los CRÍTICOS de ese documento,",
    "     y después el resto agrupado por mini-categoría.",
    "3.  La única columna que llenas es «¿Va?» (Sí / No / Ajustar / Pendiente), más «Notas / decisión».",
    "4.  Las filas rosadas, marcadas CRÍTICO, son las que se han devuelto corregidas más de una vez.",
    "5.  Los transversales salen de data/transversales.json y son los mismos en todas las pestañas.",
    "6.  El resumen de abajo se actualiza solo a medida que vas marcando.",
]):
    c = cover.cell(row=10 + i, column=1, value=t)
    c.font = Font(name=ARIAL, size=10)
    cover.merge_cells(start_row=10 + i, start_column=1, end_row=10 + i, end_column=5)

cover["A20"] = "Ejemplo de cómo llenar una fila"
cover["A20"].font = Font(name=ARIAL, size=10, bold=True)
for c, h in enumerate(["#", "Qué revisar", "¿Va?", "Notas / decisión"], start=1):
    cell = cover.cell(row=21, column=c, value=h)
    cell.font = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
    cell.fill = HDRFILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
for c, v in enumerate(["2.35", "Formato numérico gringo: 1,000.00, miles con coma.", "Sí",
                       "Sumar también el símbolo $ pegado al número."], start=1):
    cell = cover.cell(row=22, column=c, value=v)
    cell.font = Font(name=ARIAL, size=9)
    cell.fill = LEGFILL
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = BORDER
cover.row_dimensions[22].height = 30

cover["A25"] = "Resumen de la revisión"
cover["A25"].font = Font(name=ARIAL, size=12, bold=True, color=INK)
for c, h in enumerate(["Pestaña", "Ítems", "Críticos", "Sí", "No", "Ajustar",
                       "Pendiente", "Sin marcar"], start=1):
    cell = cover.cell(row=26, column=c, value=h)
    cell.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    cell.fill = HDRFILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
cover.row_dimensions[26].height = 26

for i, (hoja, n, nc, last) in enumerate(resumen):
    r = 27 + i
    q = "'%s'" % hoja
    cover.cell(row=r, column=1, value=hoja)
    cover.cell(row=r, column=2, value=n)
    cover.cell(row=r, column=3, value=nc)
    cover.cell(row=r, column=4, value='=COUNTIF(%s!F5:F%d,"Sí")' % (q, last))
    cover.cell(row=r, column=5, value='=COUNTIF(%s!F5:F%d,"No")' % (q, last))
    cover.cell(row=r, column=6, value='=COUNTIF(%s!F5:F%d,"Ajustar")' % (q, last))
    cover.cell(row=r, column=7, value='=COUNTIF(%s!F5:F%d,"Pendiente")' % (q, last))
    cover.cell(row=r, column=8, value='=B%d-SUM(D%d:G%d)' % (r, r, r))
    for c in range(1, 9):
        cell = cover.cell(row=r, column=c)
        cell.font = Font(name=ARIAL, size=10)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left" if c == 1 else "center")
        if i % 2:
            cell.fill = BANDFILL

tot = 27 + len(resumen)
cover.cell(row=tot, column=1, value="TOTAL")
for c in range(2, 9):
    L = get_column_letter(c)
    cover.cell(row=tot, column=c, value="=SUM(%s27:%s%d)" % (L, L, tot - 1))
for c in range(1, 9):
    cell = cover.cell(row=tot, column=c)
    cell.font = Font(name=ARIAL, size=10, bold=True, color=INK)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="left" if c == 1 else "center")

cover.cell(row=tot + 2, column=1,
           value="Los transversales se repiten en cada pestaña, por eso el total suma más que las "
                 "%d reglas únicas del checklist." % D["unicos"])
cover.cell(row=tot + 2, column=1).font = Font(name=ARIAL, size=9, italic=True, color="667085")
cover.merge_cells(start_row=tot + 2, start_column=1, end_row=tot + 2, end_column=8)

for c, w in enumerate([26, 10, 12, 10, 10, 11, 12, 13], start=1):
    cover.column_dimensions[get_column_letter(c)].width = w

wb.calculation.fullCalcOnLoad = True
wb.save(XLSX)

print("downloads/QA_Checklist_LATO.xlsx actualizado")
for hoja, n, nc, _ in resumen:
    print("  %-24s %3d ítems  (%d críticos)" % (hoja, n, nc))
